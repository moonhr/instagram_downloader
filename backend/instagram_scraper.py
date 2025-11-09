from openpyxl import load_workbook
import os
import re
from datetime import datetime
import time
import csv
import zipfile
import shutil
import json
import subprocess
import instaloader

def extract_shortcode(url):
    """URL에서 shortcode 추출"""
    patterns = [
        r'instagram\.com/p/([A-Za-z0-9_-]+)',
        r'instagram\.com/reel/([A-Za-z0-9_-]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None

# 전역 loader (재사용)
_loader = None

def get_instaloader():
    """Instaloader 인스턴스 가져오기"""
    global _loader
    if _loader is None:
        _loader = instaloader.Instaloader(
            download_videos=True,
            download_video_thumbnails=False,
            download_geotags=False,
            download_comments=False,
            save_metadata=False,
            compress_json=False,
            post_metadata_txt_pattern='',
            filename_pattern='{shortcode}_{typename}'
        )
        
        # 쿠키 파일이 있으면 로드
        cookie_file = 'instagram_cookies.txt'
        if os.path.exists(cookie_file):
            try:
                _loader.load_session_from_file('instagram', cookie_file)
                print("✓ 쿠키 로드 완료")
            except:
                pass
    
    return _loader

def download_with_instaloader(url, output_path, shortcode):
    """Instaloader로 다운로드 (이미지/비디오 모두 지원)"""
    try:
        print(f"  instaloader로 다운로드 시도...")
        
        loader = get_instaloader()
        post = instaloader.Post.from_shortcode(loader.context, shortcode)
        
        # 캡션 추출
        caption = post.caption if post.caption else ''
        
        # 다운로드
        loader.download_post(post, target=output_path)
        
        # 다운로드된 파일 확인
        downloaded_files = []
        for file in os.listdir(output_path):
            filepath = os.path.join(output_path, file)
            if file.startswith(shortcode):
                # .txt, .json 파일 제외
                if not (file.endswith('.txt') or file.endswith('.json')):
                    downloaded_files.append(filepath)
                elif file.endswith('.txt') or file.endswith('.json'):
                    try:
                        os.remove(filepath)
                    except:
                        pass
        
        return downloaded_files, caption
        
    except Exception as e:
        print(f"  ✗ instaloader 에러: {e}")
        return [], ''

def download_instagram_post(url, output_path):
    """gallery-dl을 사용하여 인스타그램 게시물 다운로드"""
    
    try:
        shortcode = extract_shortcode(url)
        if not shortcode:
            return {'success': False, 'error': 'Invalid Instagram URL'}
        
        print(f"  gallery-dl로 다운로드 시도...")
        
        # 쿠키 파일 확인
        cookie_file = 'instagram_cookies.txt'
        
        # gallery-dl 설정 파일 경로
        config_file = 'gallery-dl.conf'
        
        cmd = [
            'gallery-dl',
            '--config', config_file,
            '--dest', output_path,
            '--write-metadata',
            url
        ]
        
        # 쿠키 파일이 있으면 사용
        if os.path.exists(cookie_file):
            cmd.insert(1, '--cookies')
            cmd.insert(2, cookie_file)
            print(f"  쿠키 파일 사용: {cookie_file}")
        else:
            print(f"  ⚠️  쿠키 파일 없음, 로그인 없이 시도...")
        
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=60
        )
        
        downloaded_files = []
        caption = ''
        
        if result.returncode == 0 or 'HTTP/1.1" 200' in result.stderr:
            # 다운로드된 파일 확인 (모든 하위 폴더 포함)
            for root, dirs, files in os.walk(output_path):
                for file in files:
                    filepath = os.path.join(root, file)
                    if shortcode in file:
                        if file.endswith('.json'):
                            # 메타데이터에서 캡션 추출
                            try:
                                with open(filepath, 'r', encoding='utf-8') as f:
                                    metadata = json.load(f)
                                    caption = metadata.get('description', '') or metadata.get('caption', {}).get('text', '')
                                os.remove(filepath)
                            except:
                                pass
                        else:
                            # 파일을 output_path 루트로 이동
                            new_path = os.path.join(output_path, file)
                            if filepath != new_path:
                                shutil.move(filepath, new_path)
                            downloaded_files.append(new_path)
            
            if downloaded_files:
                print(f"  ✓ 다운로드 완료: {len(downloaded_files)}개 파일")
                return {
                    'success': True,
                    'caption': caption,
                    'files': downloaded_files
                }
        
        # 실패 시 에러 메시지
        error_msg = result.stderr if result.stderr else result.stdout
        print(f"  ✗ {error_msg[:200]}")
        
        if 'login required' in error_msg.lower() or '401' in error_msg:
            if not os.path.exists(cookie_file):
                return {
                    'success': False,
                    'error': '쿠키 파일이 필요합니다. python3 export_cookies.py를 실행하세요'
                }
            else:
                return {
                    'success': False,
                    'error': '쿠키가 만료되었습니다. python3 export_cookies.py를 다시 실행하세요'
                }
        
        return {
            'success': False,
            'error': '다운로드 실패'
        }
            
    except FileNotFoundError:
        return {
            'success': False,
            'error': 'gallery-dl이 설치되지 않았습니다. pip install gallery-dl 실행 후 다시 시도하세요.'
        }
    except subprocess.TimeoutExpired:
        return {
            'success': False,
            'error': '다운로드 시간 초과'
        }
    except Exception as e:
        import traceback
        print(f"상세 에러: {traceback.format_exc()}")
        return {
            'success': False,
            'error': str(e)
        }

def sanitize_filename(text):
    """파일명으로 사용 가능하도록 문자열 정리"""
    return re.sub(r'[<>:"/\\|?*]', '_', str(text))

def parse_date_to_format(date_value):
    """다양한 날짜 형식을 yy.mm.dd로 변환"""
    if isinstance(date_value, datetime):
        return date_value.strftime('%y.%m.%d')
    
    date_str = str(date_value).strip()
    
    # 다양한 날짜 형식 시도
    date_formats = [
        '%Y-%m-%d',      # 2024-01-01
        '%Y/%m/%d',      # 2024/01/01
        '%Y.%m.%d',      # 2024.01.01
        '%Y년%m월%d일',   # 2024년01월01일
        '%Y년 %m월 %d일', # 2024년 01월 01일
        '%y-%m-%d',      # 24-01-01
        '%y/%m/%d',      # 24/01/01
        '%y.%m.%d',      # 24.01.01
        '%m/%d/%Y',      # 01/01/2024
        '%d/%m/%Y',      # 01/01/2024
        '%Y%m%d',        # 20240101
    ]
    
    for fmt in date_formats:
        try:
            parsed_date = datetime.strptime(date_str, fmt)
            return parsed_date.strftime('%y.%m.%d')
        except ValueError:
            continue
    
    # 파싱 실패 시 원본 반환 (정리만)
    return sanitize_filename(date_str)

def read_file_data(file_path):
    """파일 형식에 따라 데이터 읽기 (엑셀, CSV, Numbers)"""
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.csv':
        # CSV 파일 읽기
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            rows = list(reader)
        return headers, rows
    
    elif file_ext == '.numbers':
        # Numbers 파일은 ZIP으로 압축되어 있음
        # Numbers를 CSV로 변환
        temp_dir = file_path + '_temp'
        os.makedirs(temp_dir, exist_ok=True)
        
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Numbers 파일 내부의 CSV 찾기
            csv_file = None
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file.endswith('.csv'):
                        csv_file = os.path.join(root, file)
                        break
                if csv_file:
                    break
            
            if csv_file:
                with open(csv_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames
                    rows = list(reader)
                shutil.rmtree(temp_dir)
                return headers, rows
            else:
                shutil.rmtree(temp_dir)
                raise Exception('Numbers 파일에서 데이터를 찾을 수 없습니다')
        except:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            raise Exception('Numbers 파일 형식을 지원하지 않습니다. 엑셀이나 CSV로 내보내기 해주세요')
    
    else:
        # 엑셀 파일 읽기 (.xlsx, .xls)
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            rows.append(row_dict)
        return headers, rows

def process_excel_and_download(file_path, output_dir, task_id=None, progress_status=None):
    """파일을 읽고 인스타그램 콘텐츠 다운로드"""
    try:
        # 파일 읽기
        headers, rows = read_file_data(file_path)
        
        # 필요한 컬럼 확인
        required_columns = ['Link', 'Id', 'Date']
        for col in required_columns:
            if col not in headers:
                return {
                    'success': False,
                    'message': f'파일에 "{col}" 컬럼이 없습니다. (필수: Link, Id, Date)'
                }
        
        total = len(rows)
        success_count = 0
        fail_count = 0
        failed_links = []  # 실패한 링크 저장
        
        # 진행 상황 업데이트
        if task_id and progress_status:
            progress_status[task_id]['total'] = total
        
        # 데이터 행 처리
        for idx, row in enumerate(rows, start=1):
            try:
                url = str(row['Link']).strip()
                username = str(row['Id']).strip()
                date_value = row['Date']
                
                # 날짜 형식 정리 -> yy.mm.dd
                date = parse_date_to_format(date_value)
                
                # 진행 상황 업데이트
                if task_id and progress_status:
                    progress_status[task_id].update({
                        'progress': int((idx / total) * 100),
                        'completed': idx,
                        'message': f'다운로드 중: {username} ({idx}/{total})'
                    })
                
                # 폴더 구조 생성: 인스타 아이디/날짜/
                user_dir = os.path.join(output_dir, sanitize_filename(username))
                date_dir = os.path.join(user_dir, sanitize_filename(date))
                os.makedirs(date_dir, exist_ok=True)
                
                print(f"다운로드 중: {username} - {date} - {url}")
                
                # 게시물 다운로드
                result = download_instagram_post(url, date_dir)
                
                if result['success']:
                    # 캡션(게시물 텍스트) 저장
                    caption_file = os.path.join(date_dir, '게시물.txt')
                    with open(caption_file, 'w', encoding='utf-8') as f:
                        f.write(result['caption'])
                    
                    # JSON 파일 삭제 (불필요)
                    for file in os.listdir(date_dir):
                        if file.endswith('.info.json'):
                            os.remove(os.path.join(date_dir, file))
                    
                    success_count += 1
                    print(f"✅ 다운로드 완료: {username} - {date}")
                else:
                    print(f"❌ 다운로드 실패: {result['error']}")
                    fail_count += 1
                    failed_links.append(f"{url} (ID: {username}, Date: {date})")
                
                # API 제한 방지를 위한 대기
                time.sleep(1)
                
            except Exception as e:
                print(f"❌ 다운로드 실패 (행 {idx}): {str(e)}")
                fail_count += 1
                failed_links.append(f"{url} (ID: {username}, Date: {date}) - Error: {str(e)}")
                continue
        
        # 실패한 링크가 있으면 파일로 저장
        if failed_links:
            failed_file = os.path.join(output_dir, '실패한_링크.txt')
            with open(failed_file, 'w', encoding='utf-8') as f:
                f.write(f"실패한 링크 목록 ({fail_count}개)\n")
                f.write("=" * 80 + "\n\n")
                for link in failed_links:
                    f.write(link + "\n")
            print(f"\n실패한 링크가 '{failed_file}'에 저장되었습니다.")
        
        return {
            'success': True,
            'message': f'완료: {success_count}개 성공, {fail_count}개 실패'
        }
    
    except Exception as e:
        return {
            'success': False,
            'message': f'처리 중 오류 발생: {str(e)}'
        }
